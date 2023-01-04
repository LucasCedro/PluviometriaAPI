# -*- coding: utf-8 -*-
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta, datetime

import base64
import sys, os
from pathlib import Path
import tempfile
import shutil

from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.http.request_options import RequestOptions
from office365.sharepoint.client_context import ClientContext
from office365.runtime.client_request_exception import ClientRequestException
# pip install Office365-REST-Python-Client

def downloadToTemp (file_url):
    download_path = os.path.join(tempfile.mkdtemp(), os.path.basename(file_url))
    with open(download_path, "wb") as local_file:
        file = ctx.web.get_file_by_server_relative_path(file_url).download(local_file).execute_query()
    return download_path

def uploadSharepoint (file, urlDestino):
    path = file
    with open(path, 'rb') as content_file:
        file_content = content_file.read()

    target_folder = ctx.web.get_folder_by_server_relative_path(urlDestino)
    name = os.path.basename(path)
    target_file = target_folder.upload_file(name, file_content).execute_query()
    print("2 - Uploaded: {0}".format(target_file.serverRelativeUrl))

def encrypt(clear, key):
    enc = []
    for i in range(len(clear)):
        key_c = key[i % len(key)]
        enc_c = chr((ord(clear[i]) + ord(key_c)) % 256)
        enc.append(enc_c)
    return base64.urlsafe_b64encode("".join(enc).encode()).decode()

def decrypt(enc, key):
    dec = []
    enc = base64.urlsafe_b64decode(enc).decode()
    for i in range(len(enc)):
        key_c = key[i % len(key)]
        dec_c = chr((256 + ord(enc[i]) - ord(key_c)) % 256)
        dec.append(dec_c)
    return "".join(dec)

# PLATEAU
key = 'WA69BFM@0xv4Phvx'
site_url_plateau = 'https://intrainnovatech.sharepoint.com/innovatechgestao/plateaugreen/'
login  = 'rpagestao@innovatech.com.br'
passwd = 'wr7CpsKpecK0wrbCrmRlwrDCqGnChQ=='
passwd = decrypt(passwd, key)

ctx = ClientContext(site_url_plateau).with_credentials(UserCredential(login, passwd))
web = ctx.web
ctx.load(web)
ctx.execute_query()
print("Conectando: {0}".format(web.properties['Title']))
# PLATEAU

yesterday = date.today() - timedelta(days=1)

# arquivo = pastaLeitura  + '/' + file_name
share_url = '/innovatechgestao/plateaugreen/Documentos Compartilhados/15. PLATAFORMA IRIS/02. BANCO DE DADOS/02. SILVICULTURA'
share_file = 'Base_Temperatura - TESTE.xlsx'
arquivo = downloadToTemp(share_url + '/' + share_file) #baixo do SHAREPOINT
print('1 - Download: {}'.format(arquivo))


wb = load_workbook(arquivo)
sheet = wb["Base Dados Climáticos"]  #A PLANILHA DEVE ESTAR FECHADA PARA ESTE CÓDIGO FUNCIONAR

lastrow = sheet.max_row #Ultima linha da planilha

prox_dia = datetime.strptime(sheet['N'+str(lastrow)].value,'%Y-%m-%d') + timedelta(days=1) # o primeiro dia a ser pego no range de data
prox_dia = prox_dia.strftime('%Y-%m-%d')


if not datetime.strptime(sheet['N'+str(lastrow)].value,'%Y-%m-%d').date() == yesterday:
    url = "https://apitempo.inmet.gov.br/estacao/"+prox_dia+"/"+str(yesterday)+"/S715"
    medidas_df = pd.read_json(url)
    
    dados = medidas_df.values.tolist()
    for row in dados:
        sheet.append(row)

wb.save(arquivo)
uploadSharepoint(arquivo, share_url)

#removo arquivo temporário
try:
    shutil.rmtree(os.path.dirname(arquivo))  # delete directory
except OSError as exc:
    if exc.errno != errno.ENOENT:  # ENOENT - no such file or directory
        raise  # re-raise exception
#shutil.move(src_path, dst_path)
print('3 - Removed: {}\n'.format(arquivo))
